#!/usr/bin/env python3
"""楽天 MarketSpeed II RSS 用の Excel テンプレートを生成する。

`config/weekly_holdings.yaml` の保有をそのまま並べ、現在値の列に
`=RssMarket(コード,"現在値")` を入れた状態のブックを作る。
利用者は MS2 起動中にこのファイルを開いて **保存し直すだけ** でよい。

なぜ保存が要るか: openpyxl は数式そのものではなく、Excel が計算して
保存した値を読む（`data_only=True`）。開いただけでは値が入らない。

    python scripts/make_rakuten_template.py
    python scripts/make_rakuten_template.py --output "C:/path/楽天保有.xlsx"

既存ファイルは上書きしない（`--force` で明示的に許可）。
RSS 関数を手で編集した内容を黙って壊さないため。
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_HOLDINGS = REPO_ROOT / "config" / "weekly_holdings.yaml"
DEFAULT_OUTPUT = REPO_ROOT / "output" / "楽天保有.xlsx"

HEADERS = ["銘柄名", "コード", "口座", "数量", "取得単価", "現在値", "通貨"]

# MS2 RSS は国内株が主対象。海外株・投信は現在値を空にして yfinance に任せる。
RSS_TARGET_CURRENCIES = {"JPY"}


def _rss_code(holding: dict) -> str | None:
    """RssMarket に渡す銘柄コード。国内株のみ（例: 7203.T -> 7203）。"""
    symbol = holding.get("quote_symbol")
    if not symbol or holding.get("currency") != "JPY":
        return None
    sym = str(symbol).upper()
    if not sym.endswith(".T"):
        return None
    return sym[:-2]


def build_workbook(holdings: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "保有"   # リーダーはこのシート名を優先して探す

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, h in enumerate(holdings):
        row = i + 2
        symbol = h.get("quote_symbol")
        ws.cell(row=row, column=1, value=h.get("name", ""))
        ws.cell(row=row, column=2, value=str(symbol) if symbol else "")
        ws.cell(row=row, column=3, value=h.get("account", ""))
        ws.cell(row=row, column=4, value=h.get("shares"))
        ws.cell(row=row, column=5, value=h.get("cost_price"))

        code = _rss_code(h)
        if code:
            # B列の値ではなくコード文字列を直接埋める。B列は "7203.T" 形式で
            # RssMarket が解釈できないため。
            ws.cell(row=row, column=6, value=f'=RssMarket("{code}","現在値")')
        # 海外株・投信は空欄のまま（yfinance が補完する）

        ws.cell(row=row, column=7, value=h.get("currency", ""))

    widths = {"A": 26, "B": 12, "C": 14, "D": 12, "E": 14, "F": 26, "G": 8}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    _append_notes(ws, len(holdings) + 3)
    _add_domestic_position_sheet(wb)
    return wb


#: RssPositionList の取得項目名（楽天公式オンラインヘルプの表記に一致させること）
_POSITION_LIST_COLUMNS = [
    "銘柄コード", "銘柄名称", "口座区分", "保有数量", "平均取得価額", "時価",
]


def _add_domestic_position_sheet(wb) -> None:
    """国内株の保有一覧を RssPositionList で自動取得するシートを足す。

    ``RssPositionList`` は **国内株式のみ** が対象で、米国株・投信は取れない
    （楽天公式ヘルプで確認済み）。よって「保有」シートを置き換えることはできず、
    国内株の数量・取得単価を突き合わせるための照合用シートとして持つ。

    数式は動的配列として下方向に spill するので、下に何も置かない。
    """
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("国内RSS")
    title = ws.cell(row=1, column=1,
                    value="国内株の保有一覧（MS2 が自動取得。米国株・投信は対象外）")
    title.font = Font(bold=True)

    for col, name in enumerate(_POSITION_LIST_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")

    last_col = chr(ord("A") + len(_POSITION_LIST_COLUMNS) - 1)
    # 第2引数(銘柄コード)を空にして全銘柄、第3引数 "A" で全口座区分。
    ws.cell(row=3, column=1,
            value=f'=RssPositionList($A$2:${last_col}$2,,"A")')

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 16


def _append_notes(ws, start_row: int) -> None:
    notes = [
        "■ 使い方",
        "1. MarketSpeed II を起動してログインし、RSS を有効にする",
        "2. このファイルを Excel で開く（F列の RssMarket が現在値を引く）",
        "3. 値が入ったら、そのまま上書き保存する",
        "   ※ openpyxl は数式ではなく保存済みの値を読むため、保存が必須",
        "4. config/rakuten.yaml の snapshot_path にこのファイルの絶対パスを書く",
        "",
        "■ 現在値が空欄の行について",
        "MS2 RSS は国内株が主対象。海外株・投信は空欄のままでよい。",
        "数量と取得単価だけ楽天の実データを使い、株価は yfinance が補完する。",
        "レポートの price_source 列に、どちらから来た値か必ず表示される。",
        "",
        "■ 売買したら",
        "この表の数量・取得単価を直して保存する。行の追加・削除も可。",
        "銘柄名/コード/口座/数量/取得単価/通貨 の見出しさえ保てば読める。",
        "",
        "■「国内RSS」シートについて",
        "国内株の保有一覧は MS2 が自動取得する（RssPositionList）。",
        "国内株の数量が変わったら、そちらを見てこの表を直せばよい。",
        "※ 米国株・投信は RssPositionList の対象外なので、手で直すしかない。",
    ]
    from openpyxl.styles import Font

    for i, text in enumerate(notes):
        cell = ws.cell(row=start_row + i, column=1, value=text)
        if text.startswith("■"):
            cell.font = Font(bold=True)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--holdings", default=str(DEFAULT_HOLDINGS),
                        help="保有定義 YAML のパス")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT),
                        help="生成する xlsx のパス")
    parser.add_argument("--force", action="store_true",
                        help="既存ファイルを上書きする")
    args = parser.parse_args()

    holdings_path = Path(args.holdings)
    if not holdings_path.exists():
        print(f"[error] 保有定義が見つかりません: {holdings_path}")
        return 1

    config = yaml.safe_load(holdings_path.read_text(encoding="utf-8")) or {}
    holdings = config.get("holdings") or []
    if not holdings:
        print(f"[error] {holdings_path} に holdings がありません")
        return 1

    out = Path(args.output)
    if out.exists() and not args.force:
        print(f"[error] 既に存在します: {out}\n"
              f"        上書きするなら --force を付けてください"
              f"（RSS 関数の手編集を黙って壊さないための確認です）")
        return 1

    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb = build_workbook(holdings)
    except ImportError:
        print("[error] openpyxl が必要です: pip install openpyxl")
        return 1
    wb.save(out)

    rss_rows = sum(1 for h in holdings if _rss_code(h))
    print(f"✅ 生成しました: {out}")
    print(f"   保有 {len(holdings)} 件（うち RSS で現在値を引くのは国内株 {rss_rows} 件）")
    print()
    print("次にやること:")
    print("  1. MarketSpeed II を起動してログイン（RSS 有効化）")
    print(f"  2. {out} を Excel で開く")
    print("  3. F列に現在値が入ったら、そのまま上書き保存")
    print(f'  4. config/rakuten.yaml の snapshot_path に "{out.as_posix()}" を書く')
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
