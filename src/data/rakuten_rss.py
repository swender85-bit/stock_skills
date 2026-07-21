"""楽天証券 マーケットスピードII RSS スナップショットの読み取り.

## 設計方針

証券口座への自動ログインは**しない**。無人実行のために口座資格情報を保存するのは
リスクが高すぎ、2FA も突破できず、規約上も危うい。

代わりに、あなたがログイン済みの MarketSpeed II が RSS 関数で Excel に書き出した
値を読む。資格情報はこちら側に一切来ず、それでいて楽天の実データが入る。

## 期待するファイル形式

`.xlsx`(MS2 RSS が値を書き込むブック) または `.csv`。
シート1枚目、または「保有」という名前のシートに、以下の見出し行を持つ表:

| 銘柄名 | コード | 口座 | 数量 | 取得単価 | 現在値 | 通貨 |

見出しは表記ゆれを吸収する(``銘柄`` / ``ticker`` / ``保有数`` など)。
RSS 数式の例:

    C2: =RssMarket(A2,"現在値")

MS2 RSS は国内株が主対象。米国株(SOXL/TECL/TQQQ)の現在値が取れない場合は、
数量・取得単価だけ本ファイルから採り、株価は yfinance で補完する
(``merge_with_fallback`` がその合流点)。

## 鮮度

土曜の朝に「先週の終値」を見たいので、ファイルが古いと黙って古い値を使ってしまう。
``read_snapshot`` は必ずファイル更新時刻を返し、呼び出し側が鮮度を判定できるようにする。
"""

from __future__ import annotations

import csv
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional


#: 見出しの表記ゆれ吸収
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("銘柄名", "銘柄", "名称", "商品名", "name"),
    "code": ("コード", "銘柄コード", "ティッカー", "symbol", "ticker", "code"),
    "account": ("口座", "口座区分", "account"),
    "shares": ("数量", "保有数", "保有数量", "株数", "口数", "shares", "quantity"),
    "cost_price": ("取得単価", "平均取得単価", "取得価額", "cost", "avg_price"),
    "price": ("現在値", "時価", "終値", "株価", "price", "last"),
    "currency": ("通貨", "通貨単位", "currency"),
}

_NUMERIC_FIELDS = ("shares", "cost_price", "price")


class SnapshotUnavailable(RuntimeError):
    """スナップショットが読めない(未設定・不在・空)。"""


def _normalize_header(cell: Any) -> Optional[str]:
    """見出しセルを内部フィールド名に正規化する。"""
    if cell is None:
        return None
    text = str(cell).strip().lower().replace(" ", "").replace("　", "")
    if not text:
        return None
    for field, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if text == alias.lower():
                return field
    return None


def _to_number(value: Any) -> Optional[float]:
    """カンマ・通貨記号・全角を含む数値表現を float にする。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in ("-", "—", "N/A", "#N/A"):
        return None
    for ch in (",", "¥", "$", "円", "株", "口", " ", "　"):
        text = text.replace(ch, "")
    try:
        return float(text)
    except ValueError:
        return None


def _rows_from_csv(path: Path) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            with path.open(encoding=encoding, newline="") as f:
                return [row for row in csv.reader(f)]
        except UnicodeDecodeError:
            continue
    raise SnapshotUnavailable(f"CSVの文字コードを判別できません: {path}")


def _rows_from_xlsx(path: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - 環境依存
        raise SnapshotUnavailable(
            "xlsx を読むには openpyxl が必要です: pip install openpyxl"
        ) from e

    # data_only=True で数式ではなく Excel が保存した計算済みの値を読む。
    # MS2 RSS の値は Excel 側で確定してから保存される必要がある。
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = wb["保有"] if "保有" in wb.sheetnames else wb[wb.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        wb.close()


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[int, str]]:
    """見出し行を探し、(行番号, 列index->フィールド名) を返す。"""
    for idx, row in enumerate(rows[:20]):  # 上から20行以内に見出しがある前提
        mapping = {}
        for col, cell in enumerate(row or []):
            field = _normalize_header(cell)
            if field:
                mapping[col] = field
        # 最低限、銘柄を特定する列と数量があれば表とみなす
        if "shares" in mapping.values() and (
            "code" in mapping.values() or "name" in mapping.values()
        ):
            return idx, mapping
    raise SnapshotUnavailable(
        "見出し行が見つかりません。『銘柄名 / コード / 数量 / 取得単価 / 現在値』を"
        "含むヘッダ行を用意してください。"
    )


def normalize_code(code: Any, currency: Optional[str] = None) -> str:
    """楽天の銘柄コードを yfinance ティッカーに正規化する。

    4桁の数字は日本株なので ``.T`` を付ける。既にサフィックスがあれば触らない。
    """
    if code is None:
        return ""
    text = str(code).strip().upper()
    if not text:
        return ""
    if "." in text:
        return text
    # 4桁数字 (+ 末尾英字の銘柄コードもある) は東証
    base = text.rstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    if base.isdigit() and len(text) in (4, 5):
        return f"{text}.T"
    return text


def read_snapshot(path: str | os.PathLike) -> dict:
    """MS2 RSS スナップショット(xlsx/csv)を読み込む。

    Returns
    -------
    dict
        {"holdings": [...], "source_path": str, "modified_at": ISO8601(UTC),
         "age_hours": float, "row_count": int}

    Raises
    ------
    SnapshotUnavailable
        ファイルが無い / 表が見つからない / 明細が0件。
    """
    p = Path(path)
    if not p.exists():
        raise SnapshotUnavailable(f"スナップショットが見つかりません: {p}")

    if p.suffix.lower() in (".xlsx", ".xlsm"):
        rows = _rows_from_xlsx(p)
    elif p.suffix.lower() == ".csv":
        rows = _rows_from_csv(p)
    else:
        raise SnapshotUnavailable(f"未対応の拡張子です: {p.suffix}(.xlsx / .csv)")

    header_idx, mapping = _find_header(rows)

    holdings: list[dict] = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        record: dict = {}
        for col, field in mapping.items():
            value = row[col] if col < len(row) else None
            record[field] = _to_number(value) if field in _NUMERIC_FIELDS else (
                str(value).strip() if value is not None else ""
            )

        if not record.get("shares"):
            continue  # 合計行・空行を除く

        record["quote_symbol"] = normalize_code(
            record.get("code") or "", record.get("currency")
        )
        if not record["quote_symbol"] and not record.get("name"):
            continue
        record.setdefault("currency", "")
        if not record["currency"]:
            record["currency"] = "JPY" if record["quote_symbol"].endswith(".T") else "USD"
        holdings.append(record)

    if not holdings:
        raise SnapshotUnavailable(f"明細が0件です: {p}")

    modified = datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc)
    age_hours = (datetime.now(timezone.utc) - modified).total_seconds() / 3600.0

    return {
        "holdings": holdings,
        "source_path": str(p),
        "modified_at": modified.isoformat(timespec="seconds"),
        "age_hours": round(age_hours, 1),
        "row_count": len(holdings),
    }


def merge_with_fallback(
    snapshot_holdings: list[dict],
    price_lookup: Optional[dict[str, float]] = None,
) -> list[dict]:
    """RSS で現在値が取れなかった銘柄を、別ソースの株価で補完する。

    MS2 RSS は国内株が主対象なので、米国株の現在値が空になることがある。
    数量・取得単価は楽天の実データを正、株価だけ補完する。

    Parameters
    ----------
    price_lookup : dict
        yfinance 等から取得した ``symbol -> price``。

    Returns
    -------
    list of dict
        各要素に ``price_source`` ("rakuten-rss" | "fallback" | "missing") が付く。
    """
    prices = price_lookup or {}
    merged: list[dict] = []
    for h in snapshot_holdings:
        record = dict(h)
        if record.get("price"):
            record["price_source"] = "rakuten-rss"
        else:
            fallback = prices.get(record.get("quote_symbol", ""))
            if fallback:
                record["price"] = float(fallback)
                record["price_source"] = "fallback"
            else:
                record["price_source"] = "missing"
        merged.append(record)
    return merged


def snapshot_freshness(snapshot: dict, max_age_hours: float = 72.0) -> dict:
    """スナップショットの鮮度を判定する。

    古いファイルを黙って使うと「先週の終値」のつもりで先々週の値を見る事故になる。
    """
    age = float(snapshot.get("age_hours") or 0.0)
    fresh = age <= max_age_hours
    return {
        "fresh": fresh,
        "age_hours": age,
        "max_age_hours": max_age_hours,
        "message": (
            f"MS2 RSS スナップショットは {age:.1f} 時間前のものです。"
            if fresh
            else f"⚠️ スナップショットが {age:.1f} 時間前と古いです"
            f"（許容 {max_age_hours:.0f} 時間）。MarketSpeed II で更新して保存し直してください。"
        ),
    }
